from ics import Calendar
from datetime import datetime
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import os
import iso8601


class ProcessClass:

    def __init__(self, filename):
        self.filename = filename

    def read_data(self):
        c = Calendar(open(self.filename, 'rb').read().decode('iso-8859-1'))
        c_list = list(c.events)
        if len(c_list) == 0:
            return None
        else:
            return c

    def create_event_table(self, c):
        '''
        Expects: Calendar file, read with Calendar class
        Returns: pandas dataframe, date of first event, date of last event, in iso8601
        '''
        c_list = list(c.events)
        c_list.sort()
        eventdf = pd.DataFrame(columns=['Projectnaam', 'Day', 'Duur'])

        first = iso8601.parse_date(str(c_list[0].begin))  # FIFO"%Y-%m-%dT%H:%M:%S%z")
        last = iso8601.parse_date(str(c_list[len(c_list) - 1].begin))

        for e in c.events:
            start = iso8601.parse_date(str(e.begin))
            if (
                    e.name is not None):  # extra condition: e.g. "& (start.year == 2019) & (start.month == 4)": Get year and month from user input. e.g. mail with subject 2019-05, return mail with hours
                try:
                    if ("_" in e.name):
                        end = iso8601.parse_date(str(e.end))
                        duration = end - start
                        project = e.name.split(' ')[0]
                        day = start.day  # +'/'+ str(start.month)
                        eventdf = eventdf.append(
                            {'Projectnaam': project, 'Day': day, 'Duur': duration.total_seconds() / 3600},
                            ignore_index=True)
                except:
                    print('name not iterable..')
                    continue
        return eventdf, last, first

    def create_hour_table(self, eventdf):
        '''
        Expects: pandas dataframe
        Returns: pandas dataframe
        '''
        column_names = [x for x in range(1, 32)]
        column_names = ["Projectnaam"] + column_names
        project_names = {"Projectnaam": sorted(eventdf['Projectnaam'].unique())}
        hour_table = pd.DataFrame(data=project_names, index=None, columns=column_names, dtype=None, copy=False)
        return hour_table

    def add_duration(self, eventdf, hour_table):
        for row in eventdf.itertuples():
            r = hour_table.loc[hour_table['Projectnaam'] == row.Projectnaam].index
            if pd.isnull(hour_table.iloc[r, int(row.Day)]).bool():
                hour_table.at[r, int(row.Day)] = row.Duur
            else:
                hour_table.at[r, int(row.Day)] = hour_table.iloc[r, int(row.Day)] + row.Duur
        print("Report created!")
        return hour_table

    def edit_excel(self, file_attached, first, last):
        wb = load_workbook(file_attached)
        ws = wb.active

        ws.delete_cols(1)
        ws.insert_rows(1)
        ws.insert_rows(1)

        # set column widths
        ws.column_dimensions['A'].width = 40
        for column in range(2, 33):
            column_letter = get_column_letter(column)
            ws.column_dimensions[column_letter].width = 5

        # writing to the specified cell 
        ws.cell(row=1, column=1).value = 'Activities from ' + str(first.day) + '/' + str(first.month) + '/' + str(
            first.year) + ' until ' + str(last.day) + '/' + str(last.month) + '/' + str(last.year)
        ws.cell(row=3, column=1).value = ' Project name / Day of month'

        # set the height of the row 
        ws.row_dimensions[1].height = 30

        # change font size titlerow

        from openpyxl.styles import Font
        fontStyle = Font(size="18", bold=True)
        ws.cell(row=1, column=1).font = fontStyle

        # save the file 
        wb.save(file_attached)
        wb.close()

    def delete_file(self, file_attached):
        os.remove(file_attached)
        os.remove(self.filename)
        print("Downloaded attachment and drafted report deleted!")
